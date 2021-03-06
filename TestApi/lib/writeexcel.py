import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import configparser,shutil
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,colors
from config import setting

# 读取config.ini配置文件
cf = configparser.ConfigParser()
cf.read(setting.TEST_CONFIG,encoding='UTF-8')
name = cf.get('tester','name')

class WriteExcel():
    '''文件写入数据'''
    def __init__(self,filename):
        self.filename = filename
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝文件至指定报告目录下
            shutil.copyfile(setting.SOURCE_FILE,setting.TEST_CONFIG)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write_data(self,row_n,value):
        '''
        写入测试结果
        ：param row_n：数据所在行数
        ：parnm value: 测试结果值
        ：return: 无
        '''
        font_GREEN = Font(name='宋体',color=colors.GREEN,bold=True)   
        font_RED = Font(name = '宋体',color=colors.RED,bold=True)
        font1 = Font(name = '宋体',color=colors.DARKYELLOW,bold=True)
        align = Alignment(horizontal='center',vertical='center') # 垂直水平居中

        # 获取所在行数
        L_n = 'L' + str(row_n)
        M_n = 'M' + str(row_n)


        if value == 'PASS':
            self.ws.cell(row_n,12,value)
            self.ws[L_n].font = font_GREEN
        if value == 'FAIL':
            self.ws.cell(row_n,12,value)
            self.ws[L_n].font = font_RED

        self.ws.cell(row_n,13,name)
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        self.wb.save(self.filename)


if __name__ == "__main__":
    WriteExcel(setting.SOURCE_FILE).write_data(2,'PASS')