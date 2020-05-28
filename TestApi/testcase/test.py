import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,colors
from config import setting
import configparser

cf = configparser.ConfigParser()
cf.read(setting.TEST_CONFIG,encoding='UTF-8')
name = cf.get('tester','name')


wb = load_workbook('F:\shengming\github\code-meself\TestApi\database\DemoAPITestCase.xlsx')
ws = wb.active

font_GREEN = Font(name='宋体',color=colors.GREEN,bold=True)   
font_RED = Font(name = '宋体',color=colors.RED,bold=True)
font1 = Font(name = '宋体',color=colors.DARKYELLOW,bold=True)
align = Alignment(horizontal='center',vertical='center') # 垂直水平居中

# 获取所在行数
L_n = 'L' + str(1)
M_n = 'M' + str(2)
ws.cell(20,12,'ni')
ws['L20'].font = font_RED
wb.save('F:\shengming\github\code-meself\TestApi\database\DemoAPITestCase.xlsx')